import pandas as pd
from personal.models import *
import os
from django.core.management.base import BaseCommand
from datetime import datetime
from openpyxl import load_workbook
import calendar
import xlrd
from django.db.utils import IntegrityError
class Command(BaseCommand):
    help = 'import booms'

    def add_arguments(self, parser):
        pass

    def handle(self, *args, **options):
        files = os.listdir("Ficheiros PECI")
        for file in files:
            if file.endswith("xls") or file.endswith("xlsx"):
                df = pd.read_excel("Ficheiros PECI/"+file)
                df2 = df.fillna(0)
                 
            if file.startswith("Desovas"):
                flag = False
                wb = xlrd.open_workbook("Ficheiros PECI/"+file)
                worksheet = wb.sheet_by_index(0)
                for row_index in range(worksheet.nrows):
                    cell = worksheet.cell(row_index,0)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        flag = True
                    if ( not cell.ctype == xlrd.XL_CELL_DATE and flag):
                        break
                    if flag:
                        date = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        date = str(date).split(" ")[0]
                        try:
                            models = Data.objects.get(data=date)
                        except Data.DoesNotExist:
                            models = Data(date)
                            models.save()
                        models1 = Desova(data=models,femeas = toNone(worksheet.cell(row_index,1).value),desovados = toNone(worksheet.cell(row_index,2).value),embrionados = toNone(worksheet.cell(row_index,3).value))
                        models1.save()
    
            if file.startswith("Temperaturas"):
                print(file) 
                wb = xlrd.open_workbook("Ficheiros PECI/"+file)
                ## por agora não há médias
                ano = file[13:17]
                ws = wb.sheet_by_name(str(ano))
                mes =1
                for row_index in range(ws.nrows):
                    row = ws.row(row_index)
                    if mes == 12:
                        break
                    if not isMonth(row[0].value) :
                        continue
                    mes = month_to_number(row[0].value)
                    for index in range(1,31):
                        dia =str(index)
                        if len(str(mes)) != 2:
                            mes = "0"+str(mes)
                        if len(str(dia)) != 2:
                            dia = "0"+str(dia)
                        if len (str(row[index].value)) ==0 or (str(mes) == "02" and (dia =="30" or dia =="31")):
                            continue
                        #print(str(ano)+"-"+str(mes)+"-"+str(dia))
                        try:
                            data_models = Data.objects.get(data = str(ano)+"-"+str(mes)+"-"+str(dia))
                        except Data.DoesNotExist:
                            data_models = Data(data = str(ano)+"-"+str(mes)+"-"+str(dia))
                            data_models.save()
                        models1 = Temperatura(data = data_models,temperatura=row[index].value )
                        models1.save()
                
                    ### jaula e dados
            if file.startswith("Tabela"):
                print(file)
                wb = xlrd.open_workbook("Ficheiros PECI/"+file)
                ## Eventrualmente generalizer com base no nome do ficheiro
                id_jaula = [k for k in range(20)]
                for j_id in id_jaula:    
                    worksheet = wb.sheet_by_index(0)
                    temp = {1:6,2:8,3:10,4:12,5:14,6:16,7:18,8:20,9:22}
                    peso = {5:(0,10),6:(10,15),7:(15,45),8:(45,250),9:(250,500),10:(500,None)}
                    for row in [5,6,7,8,9,10]:
                        for col in temp:
                            cell = worksheet.cell(row,col)
                            try:
                                jaula_models = Jaula.objects.get(id=j_id)    
                            except Jaula.DoesNotExist:
                                jaula_models = Jaula(id=j_id,volume = 0, massa_volumica = 0)
                                jaula_models.save()
                            models = Alimentacao(valor =cell.value,temp=temp[col],peso_inicio=peso[row][0],peso_fim=peso[row][0],id_jaula=jaula_models )
                            models.save()
        
                          
            if file.startswith("Cópia de Dados"):
                print(file) 
                wb = xlrd.open_workbook("Ficheiros PECI/"+file)
                ## Eventrualmente generalizer com base no nome do ficheiro
                worksheet = wb.sheet_by_index(1)
                table_start = False
                ano = 0
                try:
                    models = Jaula.objects.get(id = int(worksheet.name.split(" ")[1]))
                except Jaula.DoesNotExist:
                    models = Jaula(id = int(worksheet.name.split(" ")[1]),volume = 0, massa_volumica = 0)
                    models.save()
                for row_index in range(worksheet.nrows):
                    cell = worksheet.cell(row_index, 1)
                    if cell.ctype == xlrd.XL_CELL_DATE or isMonth(cell.value):
                        table_start = True
                    else:
                        if table_start:
                            break
                    if table_start:
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            date = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            date = str(date).split(" ")[0]
                            ano = date.split("-")[0]
                        else :
                            mes = month_to_number(cell.value)
                            if len(str(mes))!=2:
                                mes = "0"+str(mes)
                            date = str(ano)+"-"+str(mes)+"-"+str(15)
                        try:
                            data_models = Data.objects.get(data=date)
                        except Data.DoesNotExist:
                            data_models = Data(date)
                            data_models.save()
                        models2 = Dados(id_jaula = models, data = data_models, num_peixes = worksheet.cell(row_index,2).value,
                        PM = worksheet.cell(row_index,3).value, Biom = worksheet.cell(row_index,4).value,
                        percentagem_alimentacao = worksheet.cell(row_index,5).value, peso =worksheet.cell(row_index,6).value,
                        sacos_racao = worksheet.cell(row_index,7).value, FC = worksheet.cell(row_index,9).value,
                        peso_medio = worksheet.cell(row_index,12).value, PM_teorica_alim_real = worksheet.cell(row_index,13).value,
                        alimentacao_real = worksheet.cell(row_index,14).value, PM_teorico = worksheet.cell(row_index,15).value,
                        PM_real = worksheet.cell(row_index,16).value, percentagem_mortalidade_teorica  = worksheet.cell(row_index,18).value,
                        num_mortos_teorico = worksheet.cell(row_index,19).value, percentagem_mortalidade_real = worksheet.cell(row_index,20).value,
                        num_mortos_real = worksheet.cell(row_index,21).value, FC_real = toNone(worksheet.cell(row_index,25).value))
                        models2.save()
                ##vacina e movimentos
                #falta inserir data
                flag = False
                for row_index in range(worksheet.nrows):
                    if worksheet.cell(row_index,1).value is not None and "vacina" in str(worksheet.cell(row_index,1).value).lower() and not flag:
                        flag = True
                        continue
                    if str(worksheet.cell(row_index,1).value) ==""  and flag:
                        flag = False
                        break
                    if flag:
                        cell = worksheet.cell(row_index,2).value
                        date = xlrd.xldate_as_datetime(cell ,wb.datemode)
                        date = str(date).split(" ")[0]
                        print(date)
                        try:
                            data_models = Data.objects.get(data=date)
                        except Data.DoesNotExist:
                            data_models = Data(date)
                            data_models.save()
                        models1 = Vacina(data = data_models,id_jaula = models,num = worksheet.cell(row_index,4).value,PM = worksheet.cell(row_index,5).value)
                        models1.save()
                
                
                for row_index in range(worksheet.nrows):
                    #jaula_inicio = Jaula()
                    #jaula_fim = Jaula()
                    cell = worksheet.cell(row_index,1)
                    if worksheet.cell(row_index,1).value is not None and "tabela de entrada e saida de peixe" in str(worksheet.cell(row_index,1).value).lower():
                        flag = True
                        continue
                    if flag :
                        cell = worksheet.cell(row_index+1,1)
                        if not (cell.ctype == xlrd.XL_CELL_DATE):
                            break
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            date = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            date = str(date).split(" ")[0] 
                        
                        try:
                            data_models = Data.objects.get(data=date)
                        except Data.DoesNotExist:
                            data_models = Data(date)
                            data_models.save()
                        print(date)
                    
                        print(worksheet.cell(row_index+1,6).value)
                        try:
                            j_id = toNone(worksheet.cell(row_index+1,5).value)
                            jaula_inicio = Jaula.objects.get(id=j_id)
                        except Jaula.DoesNotExist:
                            jaula_inicio = Jaula(id=j_id,volume = 0, massa_volumica = 0)
                            jaula_inicio.save()
                        try:
                            j_id = toNone(worksheet.cell(row_index+1,6).value)
                            jaula_fim = Jaula.objects.get(id=j_id)
                        except Jaula.DoesNotExist:
                            jaula_fim = Jaula(id= j_id ,volume = 0, massa_volumica = 0)
                            jaula_fim.save()
                        mov = Movimento(data = data_models,num = worksheet.cell(row_index+1,3).value,jaula_inicio=jaula_inicio,jaula_fim=jaula_fim)
                        mov.save()
                        print("oi")
                    
          
                        
     



def dataFormat(data):
    data = str(data)
    data = data.split(" ")
    return data[0]
def dataIsValid(data):
  try:
    datetime.strptime(data, '%Y-%m-%d')
    return True
  except ValueError:
    return False

def month_to_number(month_name):
    month_name = month_name.lower()
    month_names = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    month_number = month_names.index(month_name) + 1
    return month_number
def isMonth(month_name):
    month_name = month_name.lower()
    month_names = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    return month_name in month_names

def toNone(string):
    if str(string) == '' or str(string)=='?':
        return 0
    return string
