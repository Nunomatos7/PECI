import csv
from openpyxl import load_workbook
import pandas as pd
"""wb = load_workbook(filename="../Ficheiros PECI/Cópia de Dados QDS (002).xlsx")
sheet = wb.active

csv_data =[]
for value in sheet.iter_rows(values_only=True):
    csv_data.append(list(value))

with open('Cópia de Dados QDS (002).csv','w') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data : 
        writer.writerow(line)
        """
df = pd.DataFrame(pd.read_excel("Desovas 2003-04.xlsx"))
print(df)